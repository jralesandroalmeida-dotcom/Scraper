#!/usr/bin/env python3
import asyncio, json, threading, time, io, os, glob
from flask import Flask, request, jsonify, Response, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scraper import run_scraper

app  = Flask(__name__, static_folder="static")
CORS(app)
jobs = {}

RESULTS_DIR = "sessions"
os.makedirs(RESULTS_DIR, exist_ok=True)

COLUMNS    = ["Nome do Produto","Preço Atual","Quantidade de Vendas","Vendedor",
              "Avaliações","Frete","Especificações Técnicas","Descrição Completa",
              "URLs das Imagens","URL do Produto"]
COL_WIDTHS = [48,16,22,26,24,20,52,72,60,52]

def run_job(job_id, queries, max_items, country):
    job = jobs[job_id]
    cancel_event = job["cancel"]

    def log(msg, kind="log"):
        if kind == "done":
            return
        job["events"].append(json.dumps({"type": kind, "msg": msg}))

    try:
        products = asyncio.run(
            run_scraper(queries, max_items, country, log, cancel_event=cancel_event)
        )
        job["products"] = products

        # Salva sessão em disco
        session_file = os.path.join(RESULTS_DIR, f"{job_id}.json")
        meta = {
            "id": job_id,
            "queries": queries,
            "country": country,
            "total": len(products),
            "timestamp": time.strftime("%d/%m/%Y %H:%M"),
            "products": products,
        }
        with open(session_file, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False)

        job["events"].append(json.dumps({"type": "done", "msg": json.dumps({
            "pct": 100,
            "label": f"Concluído! {len(products)} produtos.",
            "total": len(products),
        })}))
    except Exception as e:
        job["events"].append(json.dumps({"type": "err", "msg": f"Erro fatal: {e}"}))
        job["events"].append(json.dumps({"type": "done", "msg": json.dumps({"pct":0,"label":"Erro.","total":0})}))
    finally:
        job["done"] = True


def build_excel(products):
    wb = Workbook(); ws = wb.active; ws.title = "Produtos"
    hf=PatternFill("solid",fgColor="1A237E"); hfont=Font(name="Arial",bold=True,color="FFFFFF",size=11)
    ha=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cf=Font(name="Arial",size=10); aw=Alignment(vertical="top",wrap_text=True)
    at=Alignment(vertical="top"); af=PatternFill("solid",fgColor="E8EAF6")
    th=Side(style="thin",color="BDBDBD"); bd=Border(left=th,right=th,top=th,bottom=th)
    ws.row_dimensions[1].height=30
    for ci,col in enumerate(COLUMNS,1):
        c=ws.cell(1,ci,col); c.font=hfont; c.fill=hf; c.alignment=ha; c.border=bd
    wc={"Especificações Técnicas","Descrição Completa","URLs das Imagens"}
    for ri,p in enumerate(products,2):
        ws.row_dimensions[ri].height=70; fill=af if ri%2==0 else PatternFill()
        for ci,col in enumerate(COLUMNS,1):
            c=ws.cell(ri,ci,p.get(col,"")); c.font=cf; c.fill=fill; c.border=bd
            c.alignment=aw if col in wc else at
    for ci,w in enumerate(COL_WIDTHS,1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(COLUMNS))}1"
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


@app.route("/")
def index(): return send_from_directory("static","index.html")

@app.route("/api/scrape", methods=["POST"])
def start_scrape():
    d=request.json or {}
    queries=d.get("queries",[]); max_items=max(1,min(50,int(d.get("maxItems",15)))); country=d.get("country","BR")
    if not queries: return jsonify({"error":"Nenhuma busca"}), 400
    job_id=str(int(time.time()*1000))
    cancel_event = threading.Event()
    jobs[job_id]={"events":[],"done":False,"products":[],"cancel":cancel_event}
    threading.Thread(target=run_job, args=(job_id,queries,max_items,country), daemon=True).start()
    return jsonify({"jobId": job_id})

@app.route("/api/cancel/<job_id>", methods=["POST"])
def cancel_job(job_id):
    if job_id in jobs:
        jobs[job_id]["cancel"].set()
        return jsonify({"ok": True})
    return jsonify({"error": "Job não encontrado"}), 404

@app.route("/api/stream/<job_id>")
def stream(job_id):
    if job_id not in jobs: return jsonify({"error":"Não encontrado"}), 404
    def generate():
        cursor=0
        while True:
            job=jobs[job_id]
            while cursor<len(job["events"]):
                yield f"data: {job['events'][cursor]}\n\n"; cursor+=1
            if job["done"] and cursor>=len(job["events"]): break
            time.sleep(0.3)
    return Response(generate(),mimetype="text/event-stream",
                    headers={"Cache-Control":"no-cache","X-Accel-Buffering":"no"})

@app.route("/api/results/<job_id>")
def results(job_id):
    if job_id in jobs: return jsonify(jobs[job_id]["products"])
    # Tenta carregar de disco
    f = os.path.join(RESULTS_DIR, f"{job_id}.json")
    if os.path.exists(f):
        with open(f, encoding="utf-8") as fp:
            return jsonify(json.load(fp)["products"])
    return jsonify({"error":"Não encontrado"}), 404

@app.route("/api/sessions")
def list_sessions():
    files = sorted(glob.glob(os.path.join(RESULTS_DIR,"*.json")), reverse=True)
    sessions = []
    for f in files[:20]:
        try:
            with open(f, encoding="utf-8") as fp:
                d = json.load(fp)
                sessions.append({
                    "id": d["id"], "queries": d["queries"],
                    "total": d["total"], "timestamp": d["timestamp"],
                    "country": d.get("country","BR"),
                })
        except: pass
    return jsonify(sessions)

@app.route("/api/excel/<job_id>")
def excel(job_id):
    products = []
    if job_id in jobs:
        products = jobs[job_id].get("products",[])
    else:
        f = os.path.join(RESULTS_DIR, f"{job_id}.json")
        if os.path.exists(f):
            with open(f, encoding="utf-8") as fp:
                products = json.load(fp)["products"]
    if not products: return jsonify({"error":"Sem produtos"}), 404
    ts = time.strftime("%Y-%m-%d")
    return Response(build_excel(products),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename=mercadolivre_{ts}.xlsx"})

@app.route("/api/csv/<job_id>")
def csv_export(job_id):
    products = []
    if job_id in jobs:
        products = jobs[job_id].get("products",[])
    else:
        f = os.path.join(RESULTS_DIR, f"{job_id}.json")
        if os.path.exists(f):
            with open(f, encoding="utf-8") as fp:
                products = json.load(fp)["products"]
    if not products: return jsonify({"error":"Sem produtos"}), 404
    import csv, io as _io
    out = _io.StringIO()
    w = csv.DictWriter(out, fieldnames=COLUMNS)
    w.writeheader()
    for p in products:
        w.writerow({c: p.get(c,"") for c in COLUMNS})
    ts = time.strftime("%Y-%m-%d")
    return Response(out.getvalue(), mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition":f"attachment; filename=mercadolivre_{ts}.csv"})


@app.route("/api/cache/stats")
def cache_stats():
    from scraper import load_cache
    cache = load_cache()
    return jsonify({"total": len(cache), "ids": list(cache.keys())[:5]})

@app.route("/api/cache/clear", methods=["POST"])
def cache_clear():
    import os
    if os.path.exists("scraped_ids.json"):
        os.remove("scraped_ids.json")
    return jsonify({"ok": True})

if __name__ == "__main__":
    os.makedirs("static",exist_ok=True)
    print("\n"+"="*52+"\n  🛒  ML Scraper\n"+"="*52+"\n  http://localhost:5000\n")
    app.run(host="0.0.0.0",port=5000,debug=False)
